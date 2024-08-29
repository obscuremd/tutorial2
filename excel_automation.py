import openpyxl as xl
from openpyxl.chart import BarChart, Reference

import os

def new_extension_name(name):
    new_fileName = ''

    # Split the filename into name and extension
    name, extension = os.path.splitext(name)

    # Insert "2" before the extension
    new_fileName = name + '2' + extension

    return new_fileName



def add_new_chart(excel_sheet):
    wb = xl.load_workbook(excel_sheet)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)

    sheet.add_chart(chart, 'e2')

    wb.save(new_extension_name(excel_sheet))